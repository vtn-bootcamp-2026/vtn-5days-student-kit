#!/usr/bin/env python3
"""
Tạo biến thể synthetic contract cho BT3 (improve multi-format):
  - contract-001.pdf   : convert từ contract-001.docx (soffice headless)
  - contract-005-spreadsheet.xlsx : hợp đồng dạng bảng (openpyxl)

PII toàn bộ là placeholder (BR-31/SD-02): tên Nguyễn Văn A/B, email @demo.vn,
SĐT 0900.000.NNN, MST số fake. Không dùng dữ liệu thật.
"""

import os
import subprocess
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
CONTRACTS_DIR = os.path.join(HERE, "contracts")


# ---------- 1. PDF: convert contract-001.docx -> contract-001.pdf ----------
def make_pdf():
    src = os.path.join(CONTRACTS_DIR, "contract-001.docx")
    if not os.path.exists(src):
        print(f"[SKIP] Không thấy {src}", file=sys.stderr)
        return
    print("[PDF] convert contract-001.docx -> contract-001.pdf qua soffice...")
    # soffice headless convert: xuất PDF vào cùng folder
    res = subprocess.run(
        ["soffice", "--headless", "--convert-to", "pdf",
         "--outdir", CONTRACTS_DIR, src],
        capture_output=True, text=True, timeout=120,
    )
    pdf = os.path.join(CONTRACTS_DIR, "contract-001.pdf")
    if os.path.exists(pdf):
        print(f"[PDF] OK -> {pdf}")
    else:
        print(f"[PDF] FAIL. stderr: {res.stderr}", file=sys.stderr)


# ---------- 2. XLSX: contract-005-spreadsheet.xlsx ----------
def make_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "HopDong"

    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    cell_font = Font(name="Calibri", size=11)
    wrap = Alignment(wrap_text=True, vertical="top")

    # --- Phần header hợp đồng ---
    meta = [
        ("CÔNG TY TNHH VIỄN THÔNG ABC (BÊN A)", ""),
        ("Công ty TNHH Dịch vụ Mạng DEF (BÊN B)", ""),
        ("HỢP ĐỒNG CUNG CẤP DỊCH VỤ TRUYỀN DẪN SỐ 005/2026", ""),
        ("Mã hợp đồng", "HD-DEMO-005"),
        ("Ngày hiệu lực", "2026-03-01"),
        ("Ngày hết hạn", "2027-02-28"),
        ("Giá trị hợp đồng (VND)", "3.500.000.000"),
    ]
    for k, v in meta:
        ws.append([k, v])

    ws.append([])
    ws.append(["DANH MỤC ĐIỀU KHOẢN"])
    ws.append([])

    # --- Bảng điều khoản ---
    headers = ["Mục", "Điều khoản", "Nội dung", "Bằng chứng (trích nguyên văn)", "Confidence"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=ws.max_row, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = wrap
        c.border = border

    clauses = [
        ("1", "SLA", "Cam kết thời gian hoạt động 99,5%", "Bên B cam kết dịch vụ đạt mức sẵn sàng 99,5%/tháng.", 0.95),
        ("2", "Điều khoản phạt", "Phạt 8% giá trị hợp đồng nếu vi phạm SLA", "Nếu SLA dưới cam kết, Bên B bị phạt 8% giá trị hợp đồng.", 0.9),
        ("3", "Gia hạn", "Tự động gia hạn 12 tháng nếu không có thông báo", "Hợp đồng tự động gia hạn 12 tháng nếu không bên nào thông báo trước 30 ngày.", 0.92),
        ("4", "Bảo mật dữ liệu", "Không có điều khoản bảo mật dữ liệu rõ ràng", "(thiếu — không tìm thấy điều khoản bảo mật dữ liệu)", 0.3),
        ("5", "Đơn phương chấm dứt", "Bên A có thể chấm dứt bất cứ lúc nào, Bên B phải báo trước 90 ngày", "Bên A được quyền đơn phương chấm dứt; Bên B phải thông báo trước 90 ngày.", 0.88),
        ("6", "Thanh toán", "Thanh toán 30 ngày sau nhận hóa đơn", "Bên A thanh toán trong vòng 30 ngày kể từ ngày nhận hóa đơn hợp lệ.", 0.94),
    ]
    for row in clauses:
        ws.append(list(row))
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=ws.max_row, column=col)
            c.font = cell_font
            c.alignment = wrap
            c.border = border

    ws.append([])
    ws.append(["CỜ ĐỎ PHÁT HIỆN (mô phỏng)"])
    ws.append(["SLA 99,5% < 99% (cờ đỏ) · Tự động gia hạn 12 tháng (cờ đỏ) · Phạt 8% (ranh giới) · Thiếu điều khoản bảo mật (cờ đỏ)"])

    ws.append([])
    ws.append(["Người đại diện Bên A", "Nguyễn Văn A"])   # placeholder PII
    ws.append(["Email liên hệ", "nguyenvana@demo.vn"])     # placeholder
    ws.append(["SĐT liên hệ", "0900.000.005"])             # test phone
    ws.append(["MST Bên B", "0123456789 (TEST)"])          # fake

    # Column widths
    widths = [6, 22, 48, 52, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    out = os.path.join(CONTRACTS_DIR, "contract-005-spreadsheet.xlsx")
    wb.save(out)
    print(f"[XLSX] OK -> {out}")


if __name__ == "__main__":
    make_pdf()
    make_xlsx()
